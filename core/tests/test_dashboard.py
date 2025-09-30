from unittest import mock
from datetime import timedelta
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone

from openpyxl import load_workbook

from core.models import Academy, Coach
from events.models import AthleteRegistration, Event
from payments.models import Payment


class DashboardViewTests(TestCase):
    def setUp(self):
        self.login_url = reverse('admin:login')
        self.dashboard_url = reverse('core:dashboard')
        self.events_url = reverse('core:events')
        self.registrations_url = reverse('core:registrations')
        self.registrations_export_url = reverse('core:registrations-export')
        self.academy = Academy.objects.create(name='Equipe Central', city='Curitiba', state='PR')
        self.coach = Coach.objects.create(full_name='Professor Lima', academy=self.academy)
        start_at = timezone.now() + timedelta(days=10)
        deadline = start_at - timedelta(days=5)
        self.event = Event.objects.create(
            title='Open Kickboxing',
            location='Ginasio Central',
            start_at=start_at,
            registration_deadline=deadline,
            registration_fee=120,
            is_published=True,
        )
        self.staff_user = get_user_model().objects.create_user(
            username='gestor',
            password='senha-secreta',
            is_staff=True,
        )

    def test_login_obrigatorio(self):
        response = self.client.get(self.dashboard_url)
        self.assertEqual(response.status_code, 302)
        self.assertIn(self.login_url, response.url)

    def test_usuario_sem_permissao_recebe_erro(self):
        usuario = get_user_model().objects.create_user(username='atleta', password='123456')
        self.client.force_login(usuario)
        response = self.client.get(self.dashboard_url)
        self.assertEqual(response.status_code, 403)

    def test_contexto_renderiza_dados(self):
        AthleteRegistration.objects.create(
            event=self.event,
            academy=self.academy,
            coach=self.coach,
            athlete_name='Carlos Silva',
            birth_date='1998-07-15',
            practice_time=AthleteRegistration.PracticeDuration.ONE_TO_THREE,
            record_wins=4,
            record_draws=1,
            record_losses=0,
            weight_kg=70.5,
            rule_set=AthleteRegistration.RuleSet.K1_LIGHT,
            modality=AthleteRegistration.Modality.AMATEUR,
            whatsapp='41999990000',
            sex=AthleteRegistration.Sex.MALE,
            status=AthleteRegistration.Status.CONFIRMED,
        )
        self.client.force_login(self.staff_user)
        response = self.client.get(self.dashboard_url)
        self.assertEqual(response.status_code, 200)
        self.assertIn('resumo_eventos', response.context)
        self.assertEqual(response.context['resumo_eventos']['total'], 1)
        self.assertIn('inscricoes_recentes', response.context)
        self.assertGreaterEqual(len(response.context['inscricoes_recentes']), 1)

    def test_registrations_view_requires_login(self):
        response = self.client.get(self.registrations_url)
        self.assertEqual(response.status_code, 302)
        self.assertIn(self.login_url, response.url)

    def test_registrations_view_lists_entries(self):
        AthleteRegistration.objects.create(
            event=self.event,
            academy=self.academy,
            coach=self.coach,
            athlete_name='Ana Souza',
            birth_date='2001-03-22',
            practice_time=AthleteRegistration.PracticeDuration.LESS_THAN_ONE,
            record_wins=1,
            record_draws=0,
            record_losses=0,
            weight_kg=60,
            rule_set=AthleteRegistration.RuleSet.K1_LIGHT,
            modality=AthleteRegistration.Modality.AMATEUR,
            whatsapp='41988887777',
            sex=AthleteRegistration.Sex.FEMALE,
        )
        self.client.force_login(self.staff_user)
        response = self.client.get(self.registrations_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Ana Souza')
        page_obj = response.context['page_obj']
        self.assertEqual(page_obj.paginator.count, 1)

    def test_registrations_export_filters_by_modality(self):
        AthleteRegistration.objects.create(
            event=self.event,
            academy=self.academy,
            coach=self.coach,
            athlete_name='Carlos Modal',
            birth_date='1995-01-10',
            practice_time=AthleteRegistration.PracticeDuration.ONE_TO_THREE,
            record_wins=3,
            record_draws=1,
            record_losses=0,
            weight_kg=68,
            rule_set=AthleteRegistration.RuleSet.K1_LIGHT,
            modality=AthleteRegistration.Modality.AMATEUR,
            whatsapp='41999990001',
            sex=AthleteRegistration.Sex.MALE,
        )
        AthleteRegistration.objects.create(
            event=self.event,
            academy=self.academy,
            coach=self.coach,
            athlete_name='Paulo Modal Pro',
            birth_date='1996-03-18',
            practice_time=AthleteRegistration.PracticeDuration.ONE_TO_THREE,
            record_wins=5,
            record_draws=0,
            record_losses=1,
            weight_kg=72,
            rule_set=AthleteRegistration.RuleSet.K1_RULES,
            modality=AthleteRegistration.Modality.PROFESSIONAL,
            whatsapp='41999990002',
            sex=AthleteRegistration.Sex.MALE,
        )
        self.client.force_login(self.staff_user)
        response = self.client.get(self.registrations_export_url, {'modalidade': AthleteRegistration.Modality.AMATEUR})
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[1]]
        self.assertIn('Modalidade', headers)
        data_rows = list(worksheet.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(data_rows), 1)
        modality_index = headers.index('Modalidade')
        expected_label = AthleteRegistration.Modality.AMATEUR.label
        self.assertTrue(all(row[modality_index] == expected_label for row in data_rows))
        workbook.close()

    def test_events_view_lists_entries(self):
        self.client.force_login(self.staff_user)
        response = self.client.get(self.events_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Open Kickboxing')
        self.assertIn('page_obj', response.context)

    def test_registration_payment_resend_action(self):
        registration = AthleteRegistration.objects.create(
            event=self.event,
            academy=self.academy,
            coach=self.coach,
            athlete_name='Lucas Araujo',
            birth_date='2002-04-01',
            practice_time=AthleteRegistration.PracticeDuration.ONE_TO_THREE,
            record_wins=2,
            record_draws=0,
            record_losses=1,
            weight_kg=68,
            rule_set=AthleteRegistration.RuleSet.K1_LIGHT,
            modality=AthleteRegistration.Modality.AMATEUR,
            whatsapp='551199999999',
            sex=AthleteRegistration.Sex.MALE,
        )
        self.client.force_login(self.staff_user)
        payment_mock = mock.Mock(invoice_url='https://pagamento.teste/novo', bank_slip_url='')
        with mock.patch('core.views.create_payment_for_registration', return_value=payment_mock):
            response = self.client.post(
                reverse('core:registration-payment', args=[registration.pk]),
                {'action': 'resend', 'next': self.registrations_url},
                follow=True,
            )
        self.assertEqual(response.status_code, 200)
        registration.refresh_from_db()
        self.assertEqual(registration.status, AthleteRegistration.Status.PENDING)
        mensagens = list(response.context['messages'])
        self.assertTrue(any('link de pagamento' in str(msg) for msg in mensagens))

    def test_registration_payment_manual_confirmation_creates_payment(self):
        registration = AthleteRegistration.objects.create(
            event=self.event,
            academy=self.academy,
            coach=self.coach,
            athlete_name='Mariana Teles',
            birth_date='2001-09-15',
            practice_time=AthleteRegistration.PracticeDuration.ONE_TO_THREE,
            record_wins=3,
            record_draws=0,
            record_losses=0,
            weight_kg=62,
            rule_set=AthleteRegistration.RuleSet.K1_LIGHT,
            modality=AthleteRegistration.Modality.AMATEUR,
            whatsapp='551198888888',
            sex=AthleteRegistration.Sex.FEMALE,
        )
        self.client.force_login(self.staff_user)
        response = self.client.post(
            reverse('core:registration-payment', args=[registration.pk]),
            {'action': 'manual-confirm', 'next': self.registrations_url},
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        registration.refresh_from_db()
        self.assertEqual(registration.status, AthleteRegistration.Status.CONFIRMED)
        self.assertTrue(Payment.objects.filter(registration=registration).exists())
